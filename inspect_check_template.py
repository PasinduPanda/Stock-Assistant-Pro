import openpyxl
import os

path = r"D:\Antigravity\Stock Cal\check\Yesterday_Stocks_Template.xlsx"
print(f"Inspecting: {path}")

try:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 1. Inspect Headers (first 5 rows)
    print("--- Header Inspection ---")
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        print(f"Row {i+1}: {row}")

    # 2. Find SKU
    target_sku = "RS24A16881-apricotsuede-6"
    print(f"\n--- Searching for {target_sku} ---")
    found = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Convert all cell values to string to search
        row_str = [str(c).strip() if c is not None else "" for c in row]
        if target_sku in row_str:
            print(f"FOUND ROW {i+1}: {row}")
            found = True
            break

    if not found:
        print("SKU NOT FOUND in simple iteration.")

except Exception as e:
    print(f"Error: {e}")
