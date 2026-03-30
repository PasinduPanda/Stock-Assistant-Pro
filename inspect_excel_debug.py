import pandas as pd
import os
from openpyxl import load_workbook

directory = r"D:\Antigravity\Stock Cal"

print(f"Listing contents of {directory}:")
try:
    files_in_dir = os.listdir(directory)
    for f in files_in_dir:
        print(f" - '{f}'")
except Exception as e:
    print(f"Error listing dir: {e}")

# Define files again based on expected names
files = {
    "MSKU": os.path.join(directory, "MSKU Sheets.xlsx"),
    "Tiktok": os.path.join(directory, "Tiktok Template.xlsx"),
    "Warehouse": os.path.join(directory, "Warehouse sheet.xlsx")
}

print("\n--- Inspecting contents ---")

for key, path in files.items():
    print(f"\nChecking {key} at {path}...")
    if not os.path.exists(path):
        print("  -> File NOT found!")
        continue
    
    print("  -> File found.")
    try:
        # Load only the first few rows to get headers/structure
        if key == "Warehouse":
            # Warehouse might have offset headers
            df = pd.read_excel(path, usecols="A:F", nrows=10, engine='openpyxl') 
            print("  [Pandas Head (first 10 rows)]:")
            print(df.to_string())
        
        elif key == "Tiktok":
            # Tiktok might have formatting issues, use openpyxl directly to peek
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            print(f"  [OpenPyXL] Sheet Name: {ws.title}")
            print("  [First 5 Rows]:")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 5: break
                print(f"    {row}")
            wb.close()
            
        elif key == "MSKU":
             df = pd.read_excel(path, nrows=5, engine='openpyxl')
             print("  [Pandas Columns]:", df.columns.tolist())
             print(df.head(2).to_string())

    except Exception as e:
        print(f"  -> Error reading file: {e}")
