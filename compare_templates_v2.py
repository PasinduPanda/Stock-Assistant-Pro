import openpyxl
from openpyxl.styles import PatternFill
import os

# Specific Paths
good_path = r"D:\Antigravity\Stock Cal\check\06\Yesterday Stocks (3).xlsx"
bad_path = r"D:\Antigravity\Stock Cal\check\My\Yesterday_Stocks_Template.xlsx"

def get_fill_info(fill):
    if not fill: return "None"
    fg = fill.fgColor.rgb if fill.fgColor else "None"
    pattern = fill.patternType
    return f"Pattern={pattern}, FG={fg}"

def inspect(path, label):
    print(f"\n--- Inspecting {label}: {os.path.basename(path)} ---")
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active # Inspect first sheet
        
        # 1. Check Column B Style
        if 'B' in ws.column_dimensions:
            col_b = ws.column_dimensions['B']
            print(f"  Column B Fill: {get_fill_info(col_b.fill)}")
        else:
             print("  Column B has no global style.")

        # 2. Check Cell B2
        cell = ws['B2']
        if cell:
             print(f"  Cell B2 Fill: {get_fill_info(cell.fill)}")

        # 3. Conditional Formatting
        if ws.conditional_formatting:
            print(f"  Conditional Formatting Rules: {len(ws.conditional_formatting)}")
            for cf in ws.conditional_formatting:
                print(f"    Rule: {cf}")
        else:
             print("  No Conditional Formatting.")

    except Exception as e:
        print(f"  Error: {e}")

inspect(good_path, "GOOD (06)")
inspect(bad_path, "BAD (My)")
