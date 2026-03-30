import openpyxl

path = r"D:\Antigravity\Stock Cal\check\Yesterday_Stocks_Template.xlsx"
print(f"Inspecting Cell Styles for: {path}")

try:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    # Inspect first 5 data rows in Column B (Index 2)
    print("\n--- Inspecting Column B Cells ---")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, min_col=2, max_col=2)):
        cell = row[0]
        fill = cell.fill
        print(f"Cell {cell.coordinate}: Pattern={fill.patternType}, FG={fill.fgColor.rgb if fill.fgColor else 'None'}, BG={fill.bgColor.rgb if fill.bgColor else 'None'}")

    # Inspect Availability Column (C -> Index 3) just in case
    print("\n--- Inspecting Column C Cells ---")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, min_col=3, max_col=3)):
        cell = row[0]
        fill = cell.fill
        print(f"Cell {cell.coordinate}: Pattern={fill.patternType}, FG={fill.fgColor.rgb if fill.fgColor else 'None'}")

except Exception as e:
    print(f"Error: {e}")
