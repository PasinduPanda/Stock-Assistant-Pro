import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# File Paths
warehouse_file = r"D:\Antigravity\Stock Cal\Today_Stocks.xlsx"
template_file = r"D:\Antigravity\Stock Cal\Yesterday_Stocks.xlsx"
output_file = r"D:\Antigravity\Stock Cal\Updated_Stock_Template.xlsx"

# Fill Colors
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
white_fill = PatternFill(fill_type=None)

def update_stock():
    print("--- Starting Stock Update Process ---")

    # 1. Load Warehouse Data (Source)
    if not os.path.exists(warehouse_file):
        print(f"Error: Warehouse file not found: {warehouse_file}")
        return

    print(f"Loading Warehouse Data from: {os.path.basename(warehouse_file)}")
    try:
        df_warehouse = pd.read_excel(warehouse_file)
        # Create dictionary {SKU: Stock}
        # Ensure SKU is string to avoid type mismatches
        warehouse_stock = dict(zip(df_warehouse['SKU'].astype(str).str.strip(), df_warehouse['Availability']))
        print(f"Loaded {len(warehouse_stock)} SKUs from Warehouse.")
    except Exception as e:
        print(f"Failed to read warehouse file: {e}")
        return

    # 2. Load Template File (Target)
    if not os.path.exists(template_file):
        print(f"Error: Template file not found: {template_file}")
        return

    print(f"Loading Template from: {os.path.basename(template_file)}")
    try:
        wb = load_workbook(template_file)
        ws = wb.active 
    except Exception as e:
        print(f"Failed to load template file: {e}")
        return

    # Remove all tables to ensure "normal sheet" output
    if ws.tables:
        print(f"Removing {len(ws.tables)} tables from the sheet.")
        # Create a list of table names to remove to avoid modification during iteration issues
        table_names = list(ws.tables.keys())
        for name in table_names:
            del ws.tables[name]

    updated_count = 0
    highlighted_red = 0
    highlighted_yellow = 0
    highlighted_green = 0

    # 3. Iterate and Update
    # Row 1 is headers. Iterate from row 2.
    for row in ws.iter_rows(min_row=2):
        cell_sku = row[0]   # Column A
        cell_stock = row[1] # Column B
        
        sku = str(cell_sku.value).strip() if cell_sku.value else None
        
        if sku and sku in warehouse_stock:
            new_quantity = warehouse_stock[sku]
            
            if pd.isna(new_quantity):
                continue
            
            # Ensure integer
            try:
                new_quantity = int(new_quantity)
            except ValueError:
                continue

            # Update Value and format as number
            cell_stock.value = new_quantity
            cell_stock.number_format = '0' # Force Number format
            updated_count += 1
            
            # 10-7 Law Logic
            if new_quantity <= 7:
                cell_stock.fill = red_fill
                highlighted_red += 1
            elif 8 <= new_quantity <= 10:
                cell_stock.fill = yellow_fill
                highlighted_yellow += 1
            else:
                # new_quantity > 10
                cell_stock.fill = green_fill # Or white_fill if Green is only for increases? 
                # User mentioned "10-7 Law and color codings. ( That's a must )"
                # Usually 10-7 implies Low Stock warnings. >10 is safe. 
                # However, previous iteration used green. 
                # Based on "There should not be an issue for Excel sheets... 10-7 Law... Output sheet sample version"
                # If sample had green for good stock, we keep it. 
                # I'll stick to Green for > 10 to be safe and visually distinct.
                highlighted_green += 1
                
    print(f"Finished processing.")
    print(f"Total SKUs updated: {updated_count}")
    print(f"Red (<=7): {highlighted_red}")
    print(f"Yellow (8-10): {highlighted_yellow}")
    print(f"Green (>10): {highlighted_green}")

    # 4. Save
    try:
        wb.save(output_file)
        print(f"Successfully saved updated file to: {output_file}")
    except Exception as e:
        print(f"Error saving output file: {e}")

if __name__ == "__main__":
    update_stock()
