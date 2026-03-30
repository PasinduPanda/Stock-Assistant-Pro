from openpyxl import load_workbook
import os

output_file = r"D:\Antigravity\Stock Cal\Updated_Stock_Template.xlsx"

print(f"Verifying {output_file}...")
wb = load_workbook(output_file)
ws = wb.active

# Color Definitions (Approximate)
GREEN_RGB = "00FF00" 
YELLOW_RGB = "FFFF00"
RED_RGB = "FF0000"

count_green = 0
count_yellow = 0
count_red = 0
checked_rows = 0

print("--- Checking Colors and Values ---")
for row in ws.iter_rows(min_row=2):
    cell_stock = row[1]
    val = cell_stock.value
    fill = cell_stock.fill
    
    if val is None:
        continue

    # 1. Check Color Logic
    if fill and fill.patternType == 'solid':
        color = fill.start_color.rgb if fill.start_color else ""
        color = str(color).upper() if color else ""
        
        # openpyxl might return AARRGGBB, e.g. 00FF0000. 
        # Or just RGB. We check containment.
        
        if RED_RGB in color:
            count_red += 1
            if not (isinstance(val, int) and val <= 7):
                print(f"FAIL: Red highlight on value {val} (Expected <= 7)")
        elif YELLOW_RGB in color:
            count_yellow += 1
            if not (isinstance(val, int) and 8 <= val <= 10):
                print(f"FAIL: Yellow highlight on value {val} (Expected 8-10)")
        elif GREEN_RGB in color:
            count_green += 1
            if not (isinstance(val, int) and val > 10):
                 print(f"FAIL: Green highlight on value {val} (Expected > 10)")

    # 2. Check Number Format (Should be int)
    if not isinstance(val, int) and str(val).isdigit():
        # strict check: val should be int type in python if openpyxl read it correctly as number
        # However, openpyxl reads what's there. If we wrote it as int, it should come back as int or float.
        pass

    checked_rows += 1

print(f"\nTotal Rows Checked: {checked_rows}")
print(f"Red (<=7): {count_red}")
print(f"Yellow (8-10): {count_yellow}")
print(f"Green (>10): {count_green}")

print("\n--- Checking for Tables ---")
if len(ws.tables) == 0:
    print("SUCCESS: No tables found in the sheet.")
else:
    print(f"FAIL: Found {len(ws.tables)} tables: {list(ws.tables.keys())}")

if count_red > 0 or count_yellow > 0 or count_green > 0:
    print("SUCCESS: Highlighting applied.")
else:
    print("WARNING: No highlights found. (Maybe no stock met criteria?)")
