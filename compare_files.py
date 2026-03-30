import openpyxl
import os
import glob
from openpyxl.styles import PatternFill

# Directories
dir_good = r"D:\Antigravity\Stock Cal\check\06"
dir_bad = r"D:\Antigravity\Stock Cal\check\My"

def get_fill_info(fill):
    if not fill: return "None"
    fg = fill.fgColor.rgb if fill.fgColor else "None"
    pattern = fill.patternType
    return f"Pattern={pattern}, FG={fg}"

def inspect_file(path, label):
    print(f"\n--- Inspecting {label}: {os.path.basename(path)} ---")
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        # 1. Check Column B Style
        col_b = ws.column_dimensions['B']
        print(f"  Column B Fill: {get_fill_info(col_b.fill)}")
        
        # 2. Check Cell B2, B3 Style
        print("  Cell Sample (B2-B4):")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=2)):
             print(f"    {row[0].coordinate}: {get_fill_info(row[0].fill)}")

        # 3. Conditional Formatting
        print(f"  Conditional Formatting Rules: {len(ws.conditional_formatting)}")
        for cf in ws.conditional_formatting:
            print(f"    Rule: {cf}")

    except Exception as e:
        print(f"  Error: {e}")

# Find first xlsx in each
files_good = glob.glob(os.path.join(dir_good, "*.xlsx"))
files_bad = glob.glob(os.path.join(dir_bad, "*.xlsx"))

if files_good:
    inspect_file(files_good[0], "GOOD (06)")
else:
    print("No files found in 06")

if files_bad:
    inspect_file(files_bad[0], "BAD (My)")
else:
    print("No files found in My")
